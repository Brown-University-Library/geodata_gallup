# -*- coding: utf-8 -*-
"""
Pull selected gallup variables from spreadsheet to create STATA Do File
Frank Donnelly / GIS and Data Librarian / Brown University
"""

import openpyxl as xl, os
from datetime import date

thedate=date.today().strftime("%m%d%Y")
surveys={1:'gallup_covid',2:'gallup_gpss',3:'gallup_tracker',4:'gallup_world'}

rpath=os.path.join('requests','test') # MODIFY BASED ON INPUT
select_file=os.path.join(rpath,'gallup_tracker_2017_vars_TEST.xlsx') #MODIFY BASED ON INPUT
survey_file=surveys[3] #MODIFY BASED ON INPUT

dofile=os.path.join(rpath,'{}_vars_{}.do'.format(survey_file,thedate))
dtafile=os.path.join(os.path.abspath(os.getcwd()),rpath,'{}_extract_{}.dta'.format(survey_file,thedate))


#MODIFY to filter by observations - DO NOT ERASE EXAMPLES - copy, then modify
obsfilter=None
# obsfilter=None
# obsfilter='keep if inlist(STATE_NAME,"CT","MA","ME","NH","RI","VT")'
# obsfilter='keep if inrange(WP1220,18,64)'
# obsfilter='keep if SC7==2 & MONTH > 6'
# obsfilter='keep if FIPS_CODE=="44007" | FIPS_CODE=="25025"'

workbook = xl.load_workbook(select_file)
ws = workbook['Sheet1']

# May need to modify ws col and cell values based on user input
vlist=[]
for cell in ws['E']:
    if cell.value in ('x','X'): 
        vlist.append((ws.cell(row=cell.row, column=2).value))
outfile = open(dofile, "w")
outfile.writelines('keep ')
outfile.writelines(" ".join(vlist)+"\n")
if obsfilter==None:
    pass
else:
    outfile.writelines(obsfilter+"\n")
outfile.writelines('save '+dtafile+"\n")
outfile.close()
print('Created',dofile)    